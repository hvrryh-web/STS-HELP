"""
Tests for Observation Report Generator.
"""

import pytest
import numpy as np
import pandas as pd
from pathlib import Path
import tempfile
from datetime import datetime

from observation_report import (
    TailRiskAnalysis,
    ConvergenceAnalysis,
    ObservationReport,
    ObservationReportGenerator,
)


def create_sample_runs_df(n_runs: int = 100, win_rate: float = 0.5, seed: int = 42) -> pd.DataFrame:
    """Create a sample runs DataFrame for testing."""
    np.random.seed(seed)
    
    wins = np.random.random(n_runs) < win_rate
    
    data = {
        'run_index': range(n_runs),
        'seed': [seed + i for i in range(n_runs)],
        'win': wins,
        'turns': np.random.randint(5, 30, n_runs),
        'damage_taken': np.random.randint(10, 80, n_runs),
        'final_hp': np.where(wins, np.random.randint(20, 80, n_runs), 0),
        'enemy_hp': np.where(wins, 0, np.random.randint(10, 50, n_runs)),
        'cards_played': np.random.randint(10, 50, n_runs),
        'peak_strength': np.random.randint(0, 10, n_runs),
        'peak_poison': np.random.randint(0, 20, n_runs),
        'peak_orbs': np.random.randint(0, 5, n_runs),
    }
    
    return pd.DataFrame(data)


class TestTailRiskAnalysis:
    """Tests for TailRiskAnalysis dataclass."""
    
    def test_tail_risk_creation(self):
        """TailRiskAnalysis can be created."""
        tail = TailRiskAnalysis(
            percentile_5=-50.0,
            percentile_95=80.0,
            worst_case_damage=100,
            worst_case_turns=2,
            best_case_hp=75,
            tail_loss_rate=0.05,
            catastrophic_loss_count=3,
        )
        
        assert tail.percentile_5 == -50.0
        assert tail.percentile_95 == 80.0
        assert tail.worst_case_damage == 100


class TestConvergenceAnalysis:
    """Tests for ConvergenceAnalysis dataclass."""
    
    def test_convergence_analysis_creation(self):
        """ConvergenceAnalysis can be created."""
        conv = ConvergenceAnalysis(
            converged=True,
            final_win_rate=0.55,
            confidence_interval=(0.45, 0.65),
            runs_to_convergence=1000,
            variance_trend='decreasing',
            stability_score=0.9,
        )
        
        assert conv.converged is True
        assert conv.final_win_rate == 0.55
        assert conv.variance_trend == 'decreasing'


class TestObservationReportGenerator:
    """Tests for ObservationReportGenerator."""
    
    def test_generator_initialization(self):
        """Generator initializes correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            assert gen.output_dir.exists()
    
    def test_generate_observation_report(self):
        """Generator creates complete observation report."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100, win_rate=0.5)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test Scenario",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            assert report.title is not None
            assert report.scenario_name == "Test Scenario"
            assert report.character == "Ironclad"
            assert report.patch_id == "TEST-001"
            assert report.total_runs == 100
    
    def test_summary_stats_computed(self):
        """Summary statistics are computed correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100, win_rate=0.6)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            stats = report.summary_stats
            
            assert 'total_runs' in stats
            assert 'win_rate' in stats
            assert 'mean_turns' in stats
            assert 'mean_damage' in stats
            assert 0 <= stats['win_rate'] <= 1
    
    def test_tail_risk_analyzed(self):
        """Tail risk is analyzed correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            tail = report.tail_risk
            
            assert tail.percentile_5 <= tail.percentile_95
            assert tail.worst_case_damage >= 0
            assert tail.tail_loss_rate >= 0
    
    def test_convergence_analyzed(self):
        """Convergence is analyzed correctly."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            conv = report.convergence
            
            assert 0 <= conv.final_win_rate <= 1
            assert conv.confidence_interval[0] <= conv.confidence_interval[1]
            assert conv.variance_trend in ['decreasing', 'stable', 'increasing', 'unknown']
    
    def test_recommendations_generated(self):
        """Recommendations are generated."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            assert len(report.recommendations) > 0
            assert all(isinstance(r, str) for r in report.recommendations)
    
    def test_methodology_notes_generated(self):
        """Methodology notes are generated."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            assert len(report.methodology_notes) > 0
    
    def test_limitations_included(self):
        """Known limitations are included."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            assert len(report.limitations) > 0


class TestPDFGeneration:
    """Tests for PDF report generation."""
    
    def test_generate_pdf_report(self):
        """PDF report can be generated."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            pdf_path = gen.generate_pdf_report(report)
            
            assert Path(pdf_path).exists()
            assert pdf_path.endswith('.pdf')
    
    def test_pdf_contains_content(self):
        """PDF file has content."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            pdf_path = gen.generate_pdf_report(report)
            
            # Check file size > 0
            assert Path(pdf_path).stat().st_size > 0


class TestXLSXGeneration:
    """Tests for Excel report generation."""
    
    def test_generate_xlsx_report(self):
        """Excel report can be generated."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            xlsx_path = gen.generate_xlsx_report(report, df)
            
            assert Path(xlsx_path).exists()
            assert xlsx_path.endswith('.xlsx')
    
    def test_xlsx_with_raw_data(self):
        """Excel report includes raw data sample."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            xlsx_path = gen.generate_xlsx_report(report, df)
            
            # Read back and verify
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            
            assert 'Summary' in wb.sheetnames
            assert 'Detailed Statistics' in wb.sheetnames
            assert 'Raw Data Sample' in wb.sheetnames


class TestGenerateAllFormats:
    """Tests for generating all format outputs."""
    
    def test_generate_all_formats(self):
        """All formats can be generated."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=100)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Test",
                character="Ironclad",
                patch_id="TEST-001",
            )
            
            outputs = gen.generate_all_formats(report, df)
            
            assert 'pdf' in outputs
            assert 'xlsx' in outputs
            assert 'json' in outputs
            
            for fmt, path in outputs.items():
                assert Path(path).exists()


class TestEdgeCases:
    """Tests for edge cases."""
    
    def test_all_wins(self):
        """Report handles 100% win rate."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=50, win_rate=1.0)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="All Wins",
                character="Ironclad",
                patch_id="TEST-002",
            )
            
            assert report.summary_stats['win_rate'] == 1.0
            assert report.summary_stats['losses'] == 0
    
    def test_all_losses(self):
        """Report handles 0% win rate."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=50, win_rate=0.0)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="All Losses",
                character="Ironclad",
                patch_id="TEST-003",
            )
            
            assert report.summary_stats['win_rate'] == 0.0
            assert report.summary_stats['wins'] == 0
    
    def test_small_sample(self):
        """Report handles small sample size."""
        with tempfile.TemporaryDirectory() as tmpdir:
            gen = ObservationReportGenerator(output_dir=tmpdir)
            
            df = create_sample_runs_df(n_runs=10, win_rate=0.5)
            
            report = gen.generate_observation_report(
                runs_df=df,
                scenario_name="Small Sample",
                character="Ironclad",
                patch_id="TEST-004",
            )
            
            assert report.total_runs == 10
            # Confidence interval should be wide
            ci = report.summary_stats['win_rate_ci']
            assert ci[1] - ci[0] > 0.2  # Wide interval for small sample
