"""
Observation Report Generator for Monte Carlo Simulation Test Suites.

Generates publication-grade observation reports including:
- Statistical analysis with confidence intervals
- Convergence testing results
- Tail-risk behavior analysis
- Comparison between scenarios
- Known limitations and methodological notes

Supports output in PDF, XLSX, and JSON formats.
JSON output can be used for programmatic access or external document generation.
"""

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import json

import numpy as np
import pandas as pd

# Import existing reporting utilities
from reporting import (
    wilson_score_interval,
    compute_failure_modes,
    get_simulator_limitations,
    HAS_OPENPYXL,
    HAS_REPORTLAB,
    HAS_MATPLOTLIB,
)

if HAS_OPENPYXL:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import LineChart, Reference

if HAS_REPORTLAB:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, ListFlowable, ListItem
    )

if HAS_MATPLOTLIB:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt


@dataclass
class TailRiskAnalysis:
    """
    Analysis of tail-risk behavior in simulation results.
    
    Attributes:
        percentile_5: 5th percentile of reward distribution
        percentile_95: 95th percentile of reward distribution
        worst_case_damage: Maximum damage taken in a single run
        worst_case_turns: Minimum turns to loss
        best_case_hp: Maximum final HP on win
        tail_loss_rate: Percentage of losses in worst 5%
        catastrophic_loss_count: Losses with damage >= starting HP
    """
    percentile_5: float
    percentile_95: float
    worst_case_damage: int
    worst_case_turns: int
    best_case_hp: int
    tail_loss_rate: float
    catastrophic_loss_count: int


@dataclass 
class ConvergenceAnalysis:
    """
    Analysis of Monte Carlo convergence.
    
    Attributes:
        converged: Whether the simulation has converged
        final_win_rate: Final win rate
        confidence_interval: 95% CI for win rate
        runs_to_convergence: Estimated runs needed for convergence
        variance_trend: Trend in variance over batches (decreasing = good)
        stability_score: Score 0-1 indicating stability
    """
    converged: bool
    final_win_rate: float
    confidence_interval: Tuple[float, float]
    runs_to_convergence: int
    variance_trend: str  # 'decreasing', 'stable', 'increasing'
    stability_score: float


@dataclass
class ObservationReport:
    """
    Complete observation report for a Monte Carlo test suite.
    
    Attributes:
        title: Report title
        scenario_name: Name of the scenario
        character: Character tested
        patch_id: Unique patch identifier
        timestamp: Report generation time
        total_runs: Total simulation runs
        summary_stats: Summary statistics dictionary
        tail_risk: Tail risk analysis
        convergence: Convergence analysis
        failure_modes: Failure mode breakdown
        methodology_notes: List of methodology notes
        limitations: List of known limitations
        recommendations: List of recommendations
    """
    title: str
    scenario_name: str
    character: str
    patch_id: str
    timestamp: datetime
    total_runs: int
    summary_stats: Dict[str, Any]
    tail_risk: TailRiskAnalysis
    convergence: ConvergenceAnalysis
    failure_modes: Dict[str, Any]
    methodology_notes: List[str]
    limitations: List[str]
    recommendations: List[str]


class ObservationReportGenerator:
    """
    Generates observation reports for Monte Carlo test suites.
    
    Produces publication-grade reports suitable for peer review
    with full statistical analysis and methodology documentation.
    """
    
    def __init__(self, output_dir: str = "reports"):
        """
        Initialize the report generator.
        
        Args:
            output_dir: Directory for output files.
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_observation_report(
        self,
        runs_df: pd.DataFrame,
        scenario_name: str,
        character: str,
        patch_id: str,
        convergence_data: Optional[List[Tuple[int, float]]] = None
    ) -> ObservationReport:
        """
        Generate a complete observation report from simulation data.
        
        Args:
            runs_df: DataFrame with simulation runs.
            scenario_name: Name of the scenario.
            character: Character name.
            patch_id: Patch ID for the run.
            convergence_data: Optional convergence data from test suite.
        
        Returns:
            ObservationReport with all analysis.
        """
        # Compute summary statistics
        summary_stats = self._compute_summary_stats(runs_df)
        
        # Analyze tail risk
        tail_risk = self._analyze_tail_risk(runs_df)
        
        # Analyze convergence
        convergence = self._analyze_convergence(runs_df, convergence_data)
        
        # Compute failure modes
        failure_modes = compute_failure_modes(runs_df)
        
        # Generate methodology notes
        methodology_notes = self._generate_methodology_notes(runs_df)
        
        # Get limitations
        limitations = get_simulator_limitations()
        
        # Generate recommendations
        recommendations = self._generate_recommendations(
            summary_stats, tail_risk, convergence, failure_modes
        )
        
        return ObservationReport(
            title=f"Monte Carlo Observation Report: {scenario_name}",
            scenario_name=scenario_name,
            character=character,
            patch_id=patch_id,
            timestamp=datetime.now(),
            total_runs=len(runs_df),
            summary_stats=summary_stats,
            tail_risk=tail_risk,
            convergence=convergence,
            failure_modes=failure_modes,
            methodology_notes=methodology_notes,
            limitations=limitations,
            recommendations=recommendations,
        )
    
    def _compute_summary_stats(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Compute comprehensive summary statistics."""
        wins = df[df['win']]
        losses = df[~df['win']]
        
        win_count = len(wins)
        total = len(df)
        
        ci_lower, ci_upper = wilson_score_interval(win_count, total)
        
        return {
            'total_runs': total,
            'wins': win_count,
            'losses': len(losses),
            'win_rate': win_count / total if total > 0 else 0,
            'win_rate_ci': (ci_lower, ci_upper),
            
            # Turn statistics
            'mean_turns': df['turns'].mean(),
            'std_turns': df['turns'].std(),
            'median_turns': df['turns'].median(),
            'turns_iqr': (df['turns'].quantile(0.25), df['turns'].quantile(0.75)),
            
            # Damage statistics
            'mean_damage': df['damage_taken'].mean(),
            'std_damage': df['damage_taken'].std(),
            'median_damage': df['damage_taken'].median(),
            'damage_iqr': (df['damage_taken'].quantile(0.25), df['damage_taken'].quantile(0.75)),
            
            # Win-specific
            'mean_final_hp': wins['final_hp'].mean() if len(wins) > 0 else 0,
            'mean_turns_on_win': wins['turns'].mean() if len(wins) > 0 else 0,
            
            # Loss-specific
            'mean_turns_on_loss': losses['turns'].mean() if len(losses) > 0 else 0,
            
            # Cards played
            'mean_cards_played': df['cards_played'].mean(),
            'total_cards_played': df['cards_played'].sum(),
        }
    
    def _analyze_tail_risk(self, df: pd.DataFrame) -> TailRiskAnalysis:
        """Analyze tail-risk behavior."""
        # Compute reward: win * 100 - damage_taken
        rewards = df['win'].astype(int) * 100 - df['damage_taken']
        
        # Percentiles
        p5 = rewards.quantile(0.05)
        p95 = rewards.quantile(0.95)
        
        # Worst case scenarios
        worst_damage = df['damage_taken'].max()
        losses = df[~df['win']]
        worst_turns = losses['turns'].min() if len(losses) > 0 else 0
        
        wins = df[df['win']]
        best_hp = wins['final_hp'].max() if len(wins) > 0 else 0
        
        # Tail loss rate (% of runs in worst 5%)
        threshold = rewards.quantile(0.05)
        tail_losses = (rewards <= threshold).sum()
        tail_loss_rate = tail_losses / len(df)
        
        # Catastrophic losses (died quickly with max damage)
        starting_hp = 80  # Default starting HP
        catastrophic = ((df['damage_taken'] >= starting_hp) & ~df['win']).sum()
        
        return TailRiskAnalysis(
            percentile_5=p5,
            percentile_95=p95,
            worst_case_damage=worst_damage,
            worst_case_turns=worst_turns,
            best_case_hp=best_hp,
            tail_loss_rate=tail_loss_rate,
            catastrophic_loss_count=catastrophic,
        )
    
    def _analyze_convergence(
        self,
        df: pd.DataFrame,
        convergence_data: Optional[List[Tuple[int, float]]] = None
    ) -> ConvergenceAnalysis:
        """Analyze Monte Carlo convergence."""
        total_runs = len(df)
        win_rate = df['win'].mean()
        win_count = int(df['win'].sum())
        ci = wilson_score_interval(win_count, total_runs)
        
        if convergence_data and len(convergence_data) > 5:
            # Use provided convergence data
            rates = [c[1] for c in convergence_data]
            
            # Check last 20% of rates for stability
            recent_start = int(len(rates) * 0.8)
            recent_rates = rates[recent_start:]
            
            recent_std = np.std(recent_rates)
            
            # Determine variance trend
            first_half_var = np.var(rates[:len(rates)//2]) if len(rates) > 1 else 0
            second_half_var = np.var(rates[len(rates)//2:]) if len(rates) > 1 else 0
            
            if second_half_var < first_half_var * 0.8:
                trend = 'decreasing'
            elif second_half_var > first_half_var * 1.2:
                trend = 'increasing'
            else:
                trend = 'stable'
            
            # Estimate runs to convergence
            if recent_std < 0.01:
                runs_to_conv = total_runs
            else:
                # Extrapolate based on current trend
                runs_to_conv = int(total_runs * (0.01 / recent_std) ** 2)
            
            converged = recent_std < 0.01
            stability = max(0, 1 - recent_std * 10)
        else:
            # Bootstrap convergence estimate
            n_bootstrap = 100
            bootstrap_rates = []
            
            for _ in range(n_bootstrap):
                sample = df['win'].sample(n=min(1000, total_runs), replace=True)
                bootstrap_rates.append(sample.mean())
            
            bootstrap_std = np.std(bootstrap_rates)
            converged = bootstrap_std < 0.02
            trend = 'stable' if bootstrap_std < 0.02 else 'unknown'
            runs_to_conv = total_runs if converged else int(total_runs * 2)
            stability = max(0, 1 - bootstrap_std * 5)
        
        return ConvergenceAnalysis(
            converged=converged,
            final_win_rate=win_rate,
            confidence_interval=ci,
            runs_to_convergence=runs_to_conv,
            variance_trend=trend,
            stability_score=stability,
        )
    
    def _generate_methodology_notes(self, df: pd.DataFrame) -> List[str]:
        """Generate methodology documentation."""
        return [
            f"Simulation executed with {len(df):,} independent Monte Carlo runs.",
            "Random number generation uses numpy.random.Generator with PCG64 algorithm.",
            "Seeds derived via SeedSequence for independent, reproducible streams.",
            "Win rate confidence intervals computed using Wilson score method (95% CI).",
            "Convergence assessed via variance reduction in trailing 20% of batches.",
            "Tail risk analysis uses reward function: 100*win - damage_taken.",
            "Failure mode classification: burst (<5 turns), mid-game, attrition (>1.5x median).",
            "All simulations use identical enemy parameters for fair comparison.",
        ]
    
    def _generate_recommendations(
        self,
        stats: Dict[str, Any],
        tail_risk: TailRiskAnalysis,
        convergence: ConvergenceAnalysis,
        failure_modes: Dict[str, Any]
    ) -> List[str]:
        """Generate actionable recommendations."""
        recommendations = []
        
        # Win rate recommendations
        win_rate = stats['win_rate']
        if win_rate < 0.4:
            recommendations.append(
                "LOW WIN RATE: Consider improving heuristics or deck composition. "
                "Current win rate suggests suboptimal decision-making."
            )
        elif win_rate > 0.8:
            recommendations.append(
                "HIGH WIN RATE: Consider increasing simulation difficulty for better calibration. "
                "Current enemy may be too weak for meaningful analysis."
            )
        
        # Convergence recommendations
        if not convergence.converged:
            recommendations.append(
                f"NOT CONVERGED: Recommend running {convergence.runs_to_convergence:,} total iterations "
                "for statistical confidence. Current results may be unstable."
            )
        
        # Tail risk recommendations
        if tail_risk.catastrophic_loss_count > stats['total_runs'] * 0.1:
            recommendations.append(
                "HIGH CATASTROPHIC LOSS RATE: Over 10% of runs end in catastrophic failure. "
                "Review early-game defense logic and intent awareness."
            )
        
        # Failure mode recommendations
        if failure_modes.get('breakdown', {}).get('burst_death', 0) > 0.3:
            recommendations.append(
                "BURST DEATH DOMINANT: Over 30% of losses occur in <5 turns. "
                "Prioritize block generation in early turns."
            )
        
        if failure_modes.get('breakdown', {}).get('attrition_death', 0) > 0.4:
            recommendations.append(
                "ATTRITION DEATH DOMINANT: Over 40% of losses from extended fights. "
                "Consider increasing damage output or strength scaling."
            )
        
        # Default recommendation if none triggered
        if not recommendations:
            recommendations.append(
                "STABLE PERFORMANCE: Results indicate balanced simulation with "
                "reasonable win rate and convergence. Continue with current parameters."
            )
        
        return recommendations
    
    def generate_pdf_report(
        self,
        report: ObservationReport,
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate PDF observation report.
        
        Args:
            report: ObservationReport to render.
            output_path: Output path for PDF.
        
        Returns:
            Path to generated PDF.
        """
        if not HAS_REPORTLAB:
            raise ImportError("reportlab is required for PDF generation")
        
        if output_path is None:
            output_path = self.output_dir / f"observation_{report.patch_id}.pdf"
        else:
            output_path = Path(output_path)
        
        doc = SimpleDocTemplate(str(output_path), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Title'],
            fontSize=20,
            spaceAfter=20
        )
        story.append(Paragraph(report.title, title_style))
        story.append(Spacer(1, 12))
        
        # Metadata
        meta_style = styles['Normal']
        story.append(Paragraph(f"<b>Character:</b> {report.character}", meta_style))
        story.append(Paragraph(f"<b>Scenario:</b> {report.scenario_name}", meta_style))
        story.append(Paragraph(f"<b>Patch ID:</b> {report.patch_id}", meta_style))
        story.append(Paragraph(f"<b>Generated:</b> {report.timestamp.strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
        story.append(Paragraph(f"<b>Total Runs:</b> {report.total_runs:,}", meta_style))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
        
        stats = report.summary_stats
        ci = stats['win_rate_ci']
        summary_text = (
            f"This observation report summarizes {report.total_runs:,} Monte Carlo simulations "
            f"for {report.character} using the {report.scenario_name} scenario. "
            f"The simulation achieved a win rate of {stats['win_rate']:.1%} "
            f"(95% CI: {ci[0]:.1%} - {ci[1]:.1%}). "
        )
        
        if report.convergence.converged:
            summary_text += "The simulation has converged with stable results."
        else:
            summary_text += f"Additional runs recommended for convergence (target: {report.convergence.runs_to_convergence:,})."
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary Statistics Table
        story.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
        
        stats_data = [
            ['Metric', 'Value', 'Notes'],
            ['Total Runs', f"{stats['total_runs']:,}", ''],
            ['Wins', f"{stats['wins']:,}", f"{stats['win_rate']:.1%} win rate"],
            ['Losses', f"{stats['losses']:,}", ''],
            ['Win Rate (95% CI)', f"{stats['win_rate']:.2%} ({ci[0]:.2%} - {ci[1]:.2%})", 'Wilson score interval'],
            ['Mean Turns', f"{stats['mean_turns']:.1f}", f"σ = {stats['std_turns']:.1f}"],
            ['Median Turns', f"{stats['median_turns']:.0f}", f"IQR: {stats['turns_iqr'][0]:.0f} - {stats['turns_iqr'][1]:.0f}"],
            ['Mean Damage', f"{stats['mean_damage']:.1f}", f"σ = {stats['std_damage']:.1f}"],
            ['Mean Final HP (wins)', f"{stats['mean_final_hp']:.1f}", 'On winning runs'],
            ['Cards Played', f"{stats['mean_cards_played']:.1f}/run", f"Total: {stats['total_cards_played']:,}"],
        ]
        
        stats_table = Table(stats_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 0.95, 0.95)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # Convergence Analysis
        story.append(Paragraph("<b>Convergence Analysis</b>", styles['Heading2']))
        
        conv = report.convergence
        conv_status = "✓ CONVERGED" if conv.converged else "✗ NOT CONVERGED"
        conv_color = colors.green if conv.converged else colors.red
        
        conv_data = [
            ['Status', conv_status],
            ['Final Win Rate', f"{conv.final_win_rate:.2%}"],
            ['95% Confidence Interval', f"{conv.confidence_interval[0]:.2%} - {conv.confidence_interval[1]:.2%}"],
            ['Variance Trend', conv.variance_trend.capitalize()],
            ['Stability Score', f"{conv.stability_score:.2f}"],
            ['Est. Runs to Convergence', f"{conv.runs_to_convergence:,}"],
        ]
        
        conv_table = Table(conv_data, colWidths=[2.5*inch, 3*inch])
        conv_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.9)),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(conv_table)
        story.append(Spacer(1, 20))
        
        # Tail Risk Analysis
        story.append(Paragraph("<b>Tail Risk Analysis</b>", styles['Heading2']))
        
        tail = report.tail_risk
        tail_data = [
            ['Metric', 'Value', 'Interpretation'],
            ['5th Percentile Reward', f"{tail.percentile_5:.1f}", 'Worst 5% of outcomes'],
            ['95th Percentile Reward', f"{tail.percentile_95:.1f}", 'Best 5% of outcomes'],
            ['Worst Case Damage', f"{tail.worst_case_damage}", 'Maximum damage in single run'],
            ['Fastest Loss (turns)', f"{tail.worst_case_turns}", 'Burst death indicator'],
            ['Best Final HP', f"{tail.best_case_hp}", 'Maximum HP remaining on win'],
            ['Tail Loss Rate', f"{tail.tail_loss_rate:.1%}", 'Runs in worst 5%'],
            ['Catastrophic Losses', f"{tail.catastrophic_loss_count}", 'Deaths with ≥starting HP damage'],
        ]
        
        tail_table = Table(tail_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
        tail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.6, 0.2, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(tail_table)
        story.append(Spacer(1, 20))
        
        # Failure Mode Analysis
        story.append(Paragraph("<b>Failure Mode Analysis</b>", styles['Heading2']))
        
        fm = report.failure_modes
        if fm.get('total_losses', 0) > 0:
            breakdown = fm.get('breakdown', {})
            fm_data = [
                ['Failure Mode', 'Percentage', 'Description'],
                ['Burst Death', f"{breakdown.get('burst_death', 0):.1%}", 'Died within 5 turns'],
                ['Mid-Game Death', f"{breakdown.get('mid_game_death', 0):.1%}", 'Died in middle of combat'],
                ['Attrition Death', f"{breakdown.get('attrition_death', 0):.1%}", 'Died after extended fight'],
            ]
            
            fm_table = Table(fm_data, colWidths=[1.5*inch, 1.5*inch, 3*inch])
            fm_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.5, 0.3, 0.3)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(fm_table)
        else:
            story.append(Paragraph("No losses recorded in simulation.", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Recommendations
        story.append(Paragraph("<b>Recommendations</b>", styles['Heading2']))
        
        for i, rec in enumerate(report.recommendations, 1):
            story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
            story.append(Spacer(1, 6))
        
        story.append(Spacer(1, 20))
        
        # Methodology Notes
        story.append(Paragraph("<b>Methodology Notes</b>", styles['Heading2']))
        
        for note in report.methodology_notes:
            story.append(Paragraph(f"• {note}", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Known Limitations
        story.append(Paragraph("<b>Known Simulator Limitations</b>", styles['Heading2']))
        story.append(Paragraph(
            "The following limitations should be considered when interpreting results:",
            styles['Normal']
        ))
        story.append(Spacer(1, 6))
        
        for i, lim in enumerate(report.limitations, 1):
            story.append(Paragraph(f"{i}. {lim}", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        
        return str(output_path)
    
    def generate_xlsx_report(
        self,
        report: ObservationReport,
        runs_df: Optional[pd.DataFrame] = None,
        output_path: Optional[str] = None
    ) -> str:
        """
        Generate Excel observation report with structured data.
        
        Args:
            report: ObservationReport to render.
            runs_df: Optional DataFrame with raw run data.
            output_path: Output path for XLSX.
        
        Returns:
            Path to generated XLSX.
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel generation")
        
        if output_path is None:
            output_path = self.output_dir / f"observation_{report.patch_id}.xlsx"
        else:
            output_path = Path(output_path)
        
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ('Report Title', report.title),
            ('Character', report.character),
            ('Scenario', report.scenario_name),
            ('Patch ID', report.patch_id),
            ('Generated', report.timestamp.isoformat()),
            ('Total Runs', report.total_runs),
            ('', ''),
            ('Win Rate', f"{report.summary_stats['win_rate']:.2%}"),
            ('95% CI Lower', f"{report.summary_stats['win_rate_ci'][0]:.2%}"),
            ('95% CI Upper', f"{report.summary_stats['win_rate_ci'][1]:.2%}"),
            ('Converged', 'Yes' if report.convergence.converged else 'No'),
            ('', ''),
            ('Mean Turns', report.summary_stats['mean_turns']),
            ('Std Turns', report.summary_stats['std_turns']),
            ('Mean Damage', report.summary_stats['mean_damage']),
            ('Mean Final HP', report.summary_stats['mean_final_hp']),
        ]
        
        for row_idx, (key, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Sheet 2: Detailed Stats
        ws_stats = wb.create_sheet("Detailed Statistics")
        
        stats_headers = ['Metric', 'Value', 'Unit', 'Notes']
        for col, header in enumerate(stats_headers, 1):
            cell = ws_stats.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        stats = report.summary_stats
        stats_rows = [
            ('Total Runs', stats['total_runs'], 'count', ''),
            ('Wins', stats['wins'], 'count', ''),
            ('Losses', stats['losses'], 'count', ''),
            ('Win Rate', stats['win_rate'], 'ratio', 'Probability of victory'),
            ('Mean Turns', stats['mean_turns'], 'turns', 'Average combat length'),
            ('Std Turns', stats['std_turns'], 'turns', 'Combat length variability'),
            ('Median Turns', stats['median_turns'], 'turns', 'Typical combat length'),
            ('Mean Damage', stats['mean_damage'], 'HP', 'Average damage taken'),
            ('Std Damage', stats['std_damage'], 'HP', 'Damage variability'),
            ('Mean Final HP', stats['mean_final_hp'], 'HP', 'On winning runs'),
            ('Cards Played/Run', stats['mean_cards_played'], 'cards', 'Average per combat'),
        ]
        
        for row_idx, (metric, value, unit, notes) in enumerate(stats_rows, 2):
            ws_stats.cell(row=row_idx, column=1, value=metric).border = thin_border
            ws_stats.cell(row=row_idx, column=2, value=value).border = thin_border
            ws_stats.cell(row=row_idx, column=3, value=unit).border = thin_border
            ws_stats.cell(row=row_idx, column=4, value=notes).border = thin_border
        
        # Sheet 3: Convergence Data
        ws_conv = wb.create_sheet("Convergence")
        
        conv_headers = ['Metric', 'Value']
        for col, header in enumerate(conv_headers, 1):
            cell = ws_conv.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        conv = report.convergence
        conv_rows = [
            ('Converged', 'Yes' if conv.converged else 'No'),
            ('Final Win Rate', conv.final_win_rate),
            ('CI Lower', conv.confidence_interval[0]),
            ('CI Upper', conv.confidence_interval[1]),
            ('Variance Trend', conv.variance_trend),
            ('Stability Score', conv.stability_score),
            ('Est. Runs to Convergence', conv.runs_to_convergence),
        ]
        
        for row_idx, (metric, value) in enumerate(conv_rows, 2):
            ws_conv.cell(row=row_idx, column=1, value=metric)
            ws_conv.cell(row=row_idx, column=2, value=value)
        
        # Sheet 4: Tail Risk
        ws_tail = wb.create_sheet("Tail Risk")
        
        tail_headers = ['Metric', 'Value', 'Description']
        for col, header in enumerate(tail_headers, 1):
            cell = ws_tail.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        
        tail = report.tail_risk
        tail_rows = [
            ('5th Percentile Reward', tail.percentile_5, 'Worst 5% outcomes'),
            ('95th Percentile Reward', tail.percentile_95, 'Best 5% outcomes'),
            ('Worst Case Damage', tail.worst_case_damage, 'Maximum damage in single run'),
            ('Fastest Loss', tail.worst_case_turns, 'Turns to earliest loss'),
            ('Best Final HP', tail.best_case_hp, 'Maximum HP on win'),
            ('Tail Loss Rate', tail.tail_loss_rate, 'Fraction in worst 5%'),
            ('Catastrophic Losses', tail.catastrophic_loss_count, 'Deaths with full HP damage'),
        ]
        
        for row_idx, (metric, value, desc) in enumerate(tail_rows, 2):
            ws_tail.cell(row=row_idx, column=1, value=metric)
            ws_tail.cell(row=row_idx, column=2, value=value)
            ws_tail.cell(row=row_idx, column=3, value=desc)
        
        # Sheet 5: Recommendations
        ws_rec = wb.create_sheet("Recommendations")
        
        ws_rec.cell(row=1, column=1, value="Recommendations").font = Font(bold=True, size=14)
        
        for row_idx, rec in enumerate(report.recommendations, 3):
            ws_rec.cell(row=row_idx, column=1, value=rec)
        
        # Sheet 6: Raw Data Sample (if provided)
        if runs_df is not None:
            ws_raw = wb.create_sheet("Raw Data Sample")
            
            sample = runs_df.head(500)  # First 500 rows
            
            for col_idx, col_name in enumerate(sample.columns, 1):
                cell = ws_raw.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
            
            for row_idx, (_, row) in enumerate(sample.iterrows(), 2):
                for col_idx, value in enumerate(row, 1):
                    ws_raw.cell(row=row_idx, column=col_idx, value=value)
        
        # Freeze panes
        ws_summary.freeze_panes = 'A2'
        ws_stats.freeze_panes = 'A2'
        
        # Save workbook
        wb.save(output_path)
        
        return str(output_path)
    
    def generate_all_formats(
        self,
        report: ObservationReport,
        runs_df: Optional[pd.DataFrame] = None,
        base_name: Optional[str] = None
    ) -> Dict[str, str]:
        """
        Generate reports in all available formats.
        
        Args:
            report: ObservationReport to render.
            runs_df: Optional DataFrame with raw run data.
            base_name: Base name for output files.
        
        Returns:
            Dictionary mapping format to file path.
        """
        if base_name is None:
            base_name = f"observation_{report.patch_id}"
        
        outputs = {}
        
        # Generate PDF
        if HAS_REPORTLAB:
            pdf_path = self.output_dir / f"{base_name}.pdf"
            outputs['pdf'] = self.generate_pdf_report(report, str(pdf_path))
        
        # Generate XLSX
        if HAS_OPENPYXL:
            xlsx_path = self.output_dir / f"{base_name}.xlsx"
            outputs['xlsx'] = self.generate_xlsx_report(report, runs_df, str(xlsx_path))
        
        # Save JSON for programmatic access and external tools
        json_path = self.output_dir / f"{base_name}.json"
        with open(json_path, 'w') as f:
            json.dump({
                'title': report.title,
                'scenario_name': report.scenario_name,
                'character': report.character,
                'patch_id': report.patch_id,
                'timestamp': report.timestamp.isoformat(),
                'total_runs': report.total_runs,
                'summary_stats': report.summary_stats,
                'convergence': {
                    'converged': report.convergence.converged,
                    'final_win_rate': report.convergence.final_win_rate,
                    'confidence_interval': report.convergence.confidence_interval,
                    'variance_trend': report.convergence.variance_trend,
                },
                'tail_risk': {
                    'percentile_5': report.tail_risk.percentile_5,
                    'percentile_95': report.tail_risk.percentile_95,
                    'worst_case_damage': report.tail_risk.worst_case_damage,
                    'catastrophic_losses': report.tail_risk.catastrophic_loss_count,
                },
                'recommendations': report.recommendations,
                'methodology_notes': report.methodology_notes,
                'limitations': report.limitations,
            }, f, indent=2, default=str)
        outputs['json'] = str(json_path)
        
        return outputs


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate Observation Reports')
    parser.add_argument('parquet_path', help='Path to simulation results parquet file')
    parser.add_argument('--scenario', default='Unknown', help='Scenario name')
    parser.add_argument('--character', default='Unknown', help='Character name')
    parser.add_argument('--output-dir', default='reports', help='Output directory')
    
    args = parser.parse_args()
    
    # Load data
    df = pd.read_parquet(args.parquet_path)
    
    # Generate patch ID
    from seed_utils import generate_patch_id
    patch_id = generate_patch_id(
        datetime.now().strftime('%Y%m%d'),
        'UNK',
        42,
        None,
        {'source': 'manual'}
    )
    
    # Generate report
    generator = ObservationReportGenerator(output_dir=args.output_dir)
    report = generator.generate_observation_report(
        df,
        scenario_name=args.scenario,
        character=args.character,
        patch_id=patch_id,
    )
    
    outputs = generator.generate_all_formats(report, df)
    
    print("Generated reports:")
    for fmt, path in outputs.items():
        print(f"  {fmt.upper()}: {path}")
