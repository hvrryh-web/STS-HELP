"""
Reporting module for Slay the Spire simulation.
Generates Excel (XLSX) tables and PDF evaluation reports.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import numpy as np
import pandas as pd

# Excel writing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# Matplotlib for charts
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


def load_simulation_data(parquet_dir: Path) -> Dict[str, pd.DataFrame]:
    """
    Load simulation data from final parquet files.
    
    Args:
        parquet_dir: Directory containing final parquet files.
    
    Returns:
        Dictionary mapping character names to DataFrames.
    """
    final_dir = parquet_dir / "final"
    if not final_dir.exists():
        final_dir = parquet_dir
    
    data = {}
    for parquet_file in final_dir.glob("*.parquet"):
        # Extract character name from filename
        character = parquet_file.stem.split('_')[0]
        df = pd.read_parquet(parquet_file)
        data[character] = df
    
    return data


def compute_summary_stats(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Compute summary statistics for a character.
    
    Args:
        df: DataFrame with simulation results.
    
    Returns:
        Dictionary of summary statistics.
    """
    wins_df = df[df['win']]
    
    stats = {
        'runs': len(df),
        'wins': df['win'].sum(),
        'win_rate': df['win'].mean(),
        'mean_turns': df['turns'].mean(),
        'median_turns': df['turns'].median(),
        'std_turns': df['turns'].std(),
        'mean_damage': df['damage_taken'].mean(),
        'median_damage': df['damage_taken'].median(),
        'std_damage': df['damage_taken'].std(),
        'mean_final_hp': wins_df['final_hp'].mean() if len(wins_df) > 0 else 0,
        'mean_cards_played': df['cards_played'].mean(),
        'peak_poison': df['peak_poison'].max() if 'peak_poison' in df.columns else 0,
        'peak_strength': df['peak_strength'].max() if 'peak_strength' in df.columns else 0,
        'peak_orbs': df['peak_orbs'].max() if 'peak_orbs' in df.columns else 0,
    }
    
    return stats


def compute_decision_metrics(df: pd.DataFrame, lambda_param: float = 0.3, beta_param: float = 0.1) -> Dict[str, float]:
    """
    Compute decision-value metrics (EV, PV, RV, APV, UPV, etc.).
    
    Args:
        df: DataFrame with simulation results.
        lambda_param: APV adjustment parameter.
        beta_param: CGV risk penalty parameter.
    
    Returns:
        Dictionary of decision metrics.
    """
    # Compute reward: win*100 - damage_taken
    df = df.copy()
    df['reward'] = df['win'].astype(int) * 100 - df['damage_taken']
    
    # RV (Return Value): mean observed reward
    rv = df['reward'].mean()
    
    # PV (Prediction Value): heuristic baseline (placeholder)
    pv = 50.0  # Default prediction
    
    # APV (Adjusted Predictive Value)
    apv = pv + lambda_param * (rv - pv)
    
    # UPV (Updated Prediction Value) - same as APV for single batch
    upv = apv
    
    # GGV (Greed God Value): 95th percentile of reward
    ggv = df['reward'].quantile(0.95)
    
    # SGV (Scared God Value): negative of 5th percentile
    sgv = -df['reward'].quantile(0.05)
    
    # CGV (Content God Value): APV penalized by variance
    cgv = apv - beta_param * df['reward'].std()
    
    # ATV (Ambitious Transcendent Value): APV * win probability
    win_prob = df['win'].mean()
    atv = apv * win_prob
    
    # JV (Jackpot Value): probability of high reward * expected high reward
    high_threshold = df['reward'].quantile(0.9)
    high_reward_mask = df['reward'] >= high_threshold
    jv = high_reward_mask.mean() * df[high_reward_mask]['reward'].mean() if high_reward_mask.any() else 0
    
    # EV (Estimated Value): baseline expected contribution
    ev = rv
    
    # NPV (Negative Predictive Value): expected downside
    losses = df[df['reward'] < 0]['reward']
    npv = -losses.mean() if len(losses) > 0 else 0
    
    return {
        'EV': ev,
        'PV': pv,
        'RV': rv,
        'NPV': npv,
        'APV': apv,
        'UPV': upv,
        'GGV': ggv,
        'SGV': sgv,
        'CGV': cgv,
        'ATV': atv,
        'JV': jv
    }


def generate_excel(
    parquet_dir: str,
    patch_id: str,
    output_path: Optional[str] = None
) -> str:
    """
    Generate Excel report from simulation data.
    
    Args:
        parquet_dir: Directory containing parquet files.
        patch_id: Patch ID for the report.
        output_path: Output path for Excel file.
    
    Returns:
        Path to generated Excel file.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel generation")
    
    parquet_path = Path(parquet_dir)
    data = load_simulation_data(parquet_path)
    
    if not data:
        raise ValueError(f"No simulation data found in {parquet_dir}")
    
    # Create workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_headers = [
        'Character', 'Relic', 'Runs', 'Win Rate', 'Mean Turns', 
        'Mean Damage', 'Std Damage', 'Mean Final HP', 'Peak Metric'
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for character, df in data.items():
        stats = compute_summary_stats(df)
        relic = df['relic'].iloc[0] if 'relic' in df.columns else 'none'
        
        peak_metric = max(stats['peak_poison'], stats['peak_strength'], stats['peak_orbs'])
        
        values = [
            character, relic, stats['runs'], f"{stats['win_rate']:.2%}",
            f"{stats['mean_turns']:.1f}", f"{stats['mean_damage']:.1f}",
            f"{stats['std_damage']:.1f}", f"{stats['mean_final_hp']:.1f}",
            peak_metric
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            cell.border = thin_border
        
        row += 1
    
    # Sheet 2: Aggregated Metrics
    ws_metrics = wb.create_sheet("Aggregated Metrics")
    
    metric_headers = ['Character', 'Relic', 'EV', 'PV', 'RV', 'NPV', 'APV', 'UPV', 'GGV', 'SGV', 'CGV', 'ATV', 'JV']
    
    for col, header in enumerate(metric_headers, 1):
        cell = ws_metrics.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for character, df in data.items():
        metrics = compute_decision_metrics(df)
        relic = df['relic'].iloc[0] if 'relic' in df.columns else 'none'
        
        values = [character, relic] + [f"{metrics[k]:.2f}" for k in ['EV', 'PV', 'RV', 'NPV', 'APV', 'UPV', 'GGV', 'SGV', 'CGV', 'ATV', 'JV']]
        
        for col, value in enumerate(values, 1):
            cell = ws_metrics.cell(row=row, column=col, value=value)
            cell.border = thin_border
        
        row += 1
    
    # Sheet 3: Deck Composition (template)
    ws_deck = wb.create_sheet("Deck Composition")
    
    deck_headers = ['Character', 'Slot', 'Card Type', 'Count (Starter)', 'Count (Optimized)', 'Notes']
    
    for col, header in enumerate(deck_headers, 1):
        cell = ws_deck.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Add Ironclad deck template
    ironclad_deck = [
        ('Ironclad', 'A', 'Strike (basic)', 5, '0-2', 'Remove as upgrades arrive'),
        ('Ironclad', 'B', 'Defend (basic)', 4, '0-2', 'Keep minimal for early survival'),
        ('Ironclad', 'C', 'Inflame / Strength', 0, '1-2', 'Core scaling'),
        ('Ironclad', 'D', 'Bash / Vulnerable', 1, '1', 'Amplify heavy attacks'),
        ('Ironclad', 'E', 'Heavy Attack', 0, '1-2', 'High payoff cards'),
        ('Ironclad', 'F', 'Block cards', 0, '2-4', 'Intent-aware defense'),
    ]
    
    for row_idx, deck_row in enumerate(ironclad_deck, 2):
        for col, value in enumerate(deck_row, 1):
            cell = ws_deck.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Sheet 4: Raw Sample
    ws_raw = wb.create_sheet("Raw Sample")
    
    # Take first 100 rows from first character as sample
    first_char = list(data.keys())[0]
    sample_df = data[first_char].head(100)
    
    for col, header in enumerate(sample_df.columns, 1):
        cell = ws_raw.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row_idx, (_, data_row) in enumerate(sample_df.iterrows(), 2):
        for col, value in enumerate(data_row, 1):
            cell = ws_raw.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Sheet 5: Patch Log
    ws_patch = wb.create_sheet("Patch Log")
    
    patch_data = [
        ('Patch ID', patch_id),
        ('Generated', datetime.now().isoformat()),
        ('Characters', ', '.join(data.keys())),
        ('Total Runs', sum(len(df) for df in data.values())),
    ]
    
    for row_idx, (key, value) in enumerate(patch_data, 1):
        ws_patch.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws_patch.cell(row=row_idx, column=2, value=str(value))
    
    # Freeze panes
    ws_summary.freeze_panes = 'A2'
    ws_metrics.freeze_panes = 'A2'
    
    # Save workbook
    if output_path is None:
        output_path = parquet_path / f"report_{patch_id}.xlsx"
    else:
        output_path = Path(output_path)
    
    wb.save(output_path)
    
    return str(output_path)


def generate_charts(data: Dict[str, pd.DataFrame], output_dir: Path) -> List[str]:
    """
    Generate chart images for PDF report.
    
    Args:
        data: Dictionary of character DataFrames.
        output_dir: Directory to save chart images.
    
    Returns:
        List of chart image paths.
    """
    if not HAS_MATPLOTLIB:
        return []
    
    charts = []
    
    # Win rate comparison chart
    fig, ax = plt.subplots(figsize=(8, 5))
    
    characters = list(data.keys())
    win_rates = [data[c]['win'].mean() * 100 for c in characters]
    
    bars = ax.bar(characters, win_rates, color=['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000'][:len(characters)])
    ax.set_ylabel('Win Rate (%)')
    ax.set_title('Win Rate by Character')
    ax.set_ylim(0, 100)
    
    for bar, rate in zip(bars, win_rates):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1, 
                f'{rate:.1f}%', ha='center', va='bottom')
    
    chart_path = output_dir / 'win_rate_chart.png'
    plt.tight_layout()
    plt.savefig(chart_path, dpi=150)
    plt.close()
    charts.append(str(chart_path))
    
    # Damage distribution chart
    fig, ax = plt.subplots(figsize=(8, 5))
    
    for i, (character, df) in enumerate(data.items()):
        ax.hist(df['damage_taken'], bins=30, alpha=0.6, label=character)
    
    ax.set_xlabel('Damage Taken')
    ax.set_ylabel('Frequency')
    ax.set_title('Damage Distribution by Character')
    ax.legend()
    
    chart_path = output_dir / 'damage_distribution.png'
    plt.tight_layout()
    plt.savefig(chart_path, dpi=150)
    plt.close()
    charts.append(str(chart_path))
    
    return charts


def generate_pdf(
    parquet_dir: str,
    patch_id: str,
    output_path: Optional[str] = None
) -> str:
    """
    Generate PDF evaluation report from simulation data.
    
    Args:
        parquet_dir: Directory containing parquet files.
        patch_id: Patch ID for the report.
        output_path: Output path for PDF file.
    
    Returns:
        Path to generated PDF file.
    """
    if not HAS_REPORTLAB:
        raise ImportError("reportlab is required for PDF generation")
    
    parquet_path = Path(parquet_dir)
    data = load_simulation_data(parquet_path)
    
    if not data:
        raise ValueError(f"No simulation data found in {parquet_dir}")
    
    # Determine output path
    if output_path is None:
        output_path = parquet_path / f"report_{patch_id}.pdf"
    else:
        output_path = Path(output_path)
    
    # Generate charts
    chart_paths = []
    if HAS_MATPLOTLIB:
        chart_paths = generate_charts(data, parquet_path)
    
    # Create PDF document
    doc = SimpleDocTemplate(str(output_path), pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Title'],
        fontSize=24,
        spaceAfter=20
    )
    story.append(Paragraph("Slay the Spire Simulation Report", title_style))
    story.append(Spacer(1, 12))
    
    # Patch ID and parameters
    story.append(Paragraph(f"<b>Patch ID:</b> {patch_id}", styles['Normal']))
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Paragraph(f"<b>Characters:</b> {', '.join(data.keys())}", styles['Normal']))
    story.append(Paragraph(f"<b>Total Runs:</b> {sum(len(df) for df in data.values())}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Executive Summary
    story.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
    
    summary_text = "This report presents the results of combat simulations for Slay the Spire characters. "
    summary_text += "The simulations use deterministic RNG for reproducibility and measure win rates, "
    summary_text += "damage taken, turns to victory, and peak scaling metrics (poison, strength, orbs). "
    summary_text += "Results can be used for deck optimization and pathing decisions."
    story.append(Paragraph(summary_text, styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary statistics table
    story.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
    
    table_data = [['Character', 'Runs', 'Win Rate', 'Mean Turns', 'Mean Damage', 'Mean Final HP']]
    
    for character, df in data.items():
        stats = compute_summary_stats(df)
        table_data.append([
            character,
            str(stats['runs']),
            f"{stats['win_rate']:.1%}",
            f"{stats['mean_turns']:.1f}",
            f"{stats['mean_damage']:.1f}",
            f"{stats['mean_final_hp']:.1f}"
        ])
    
    table = Table(table_data, colWidths=[1.2*inch, 0.8*inch, 1*inch, 1*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Decision Metrics table
    story.append(Paragraph("<b>Decision Value Metrics</b>", styles['Heading2']))
    
    metrics_data = [['Character', 'EV', 'APV', 'GGV', 'SGV', 'CGV', 'ATV', 'JV']]
    
    for character, df in data.items():
        metrics = compute_decision_metrics(df)
        metrics_data.append([
            character,
            f"{metrics['EV']:.1f}",
            f"{metrics['APV']:.1f}",
            f"{metrics['GGV']:.1f}",
            f"{metrics['SGV']:.1f}",
            f"{metrics['CGV']:.1f}",
            f"{metrics['ATV']:.1f}",
            f"{metrics['JV']:.1f}"
        ])
    
    metrics_table = Table(metrics_data, colWidths=[1.2*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.27, 0.45, 0.77)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 20))
    
    # Charts
    if chart_paths:
        story.append(Paragraph("<b>Visualizations</b>", styles['Heading2']))
        for chart_path in chart_paths:
            if os.path.exists(chart_path):
                img = Image(chart_path, width=5*inch, height=3.5*inch)
                story.append(img)
                story.append(Spacer(1, 12))
    
    # Observations and recommendations
    story.append(Paragraph("<b>Observations</b>", styles['Heading2']))
    
    observations = []
    for character, df in data.items():
        stats = compute_summary_stats(df)
        win_rate = stats['win_rate']
        
        if win_rate > 0.7:
            observations.append(f"• {character}: Strong performance ({win_rate:.1%} win rate). Consider increasing enemy difficulty for calibration.")
        elif win_rate < 0.4:
            observations.append(f"• {character}: Low win rate ({win_rate:.1%}). Consider deck improvements or reviewing heuristics.")
        else:
            observations.append(f"• {character}: Baseline performance ({win_rate:.1%} win rate). Suitable for calibration anchor.")
    
    for obs in observations:
        story.append(Paragraph(obs, styles['Normal']))
    
    story.append(Spacer(1, 12))
    
    # Next steps
    story.append(Paragraph("<b>Recommended Next Steps</b>", styles['Heading2']))
    next_steps = [
        "1. Compare results against community ground truth data for calibration",
        "2. Adjust λ and β parameters based on observed APV bias",
        "3. Run extended simulations (10k+ runs) for statistical significance",
        "4. Implement additional card and relic effects for higher fidelity"
    ]
    for step in next_steps:
        story.append(Paragraph(step, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    
    return str(output_path)


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python reporting.py <parquet_dir> <patch_id>")
        sys.exit(1)
    
    parquet_dir = sys.argv[1]
    patch_id = sys.argv[2]
    
    print(f"Generating Excel report...")
    excel_path = generate_excel(parquet_dir, patch_id)
    print(f"  Created: {excel_path}")
    
    print(f"Generating PDF report...")
    pdf_path = generate_pdf(parquet_dir, patch_id)
    print(f"  Created: {pdf_path}")
